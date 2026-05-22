#!/usr/bin/env python3
"""
Extrator de Insumos - Web App
Execute: python app.py
Abra:    http://localhost:5000
"""

import os, sys, json, base64, re, time, threading, uuid, tempfile, shutil
from pathlib import Path
from datetime import datetime, timedelta

from flask import Flask, request, jsonify, send_file
from werkzeug.utils import secure_filename
from google import genai
from google.genai import types
import fitz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 500 * 1024 * 1024  # 500 MB

JOBS: dict = {}
JOBS_LOCK = threading.Lock()
TEMP_DIR = Path(tempfile.mkdtemp(prefix="insumos_"))

# ---------------------------------------------------------------------------
# Categorias
# ---------------------------------------------------------------------------
CATEGORIAS = [
    "EQUIPE EMPREITA", "EQUIPAMENTO EMPREITA", "COMBUSTÍVEL", "CONCRETO",
    "FÔRMA", "AÇO", "M.O.", "CIMENTO", "AGLOMERANTE", "CERÂMICA",
    "PINTURA", "ALVENARIA", "AGREGADOS", "LONA", "PAVIMENTAÇÃO",
    "PAISAGISMO", "ARTEFATOS DE CONCRETO PARA PV", "TAMPÃO EM FD",
    "CADASTRO", "TOPOGRAFIA", "TERRAPLANAGEM", "ROCHA", "GEOGRELHA",
    "PVC", "GEOTEXTIL", "MOTOBOMBA", "GEOFORMAS", "CANTEIRO", "HÉLICE",
    "ESTACAS", "TUBULÃO", "ESPAÇADOR", "ALO", "SEGURANÇA DO TRABALHO",
    "COBERTURA", "FOSSAS", "MONTAGENS", "FRETE", "DESCARTE DE ENTULHO",
    "ALVARÁ", "DISCO DE CORTE", "OBRA CIVIL", "ETE MODULAR", "ELÉTRICA",
    "SERVIÇO", "URBANIZAÇÃO", "TUBO DE CONCRETO", "IMPERMEABILIZAÇÃO",
    "PRÉ OPERAÇÃO", "LINEARES", "CALHA", "ARTEFATO DE CONCRETO", "MURO",
    "PRÉ-MOLDADO", "LAJOTA", "PROJETO", "TESTES", "DRENAGEM", "FD",
    "VALVULA", "AÇO CARBONO", "COMPORTA", "PEAD", "FG", "CAÇAMBA",
    "PRFV", "MEDIDOR DE VAZÃO", "TALHA", "ESQUADRIAS", "ACESSORIO FIXAÇÃO",
    "MOVEIS", "WETLAND", "CASA SOPRADOR", "CASA QUIMICA", "INTERLIGAÇÃO",
]

CATEGORIAS_LISTA = "\n".join(f"  - {c}" for c in CATEGORIAS)

EXTRACTION_PROMPT = f"""Você é um especialista em leitura de pranchas de projetos de engenharia hidráulica.
Analise a imagem da prancha e extraia as informações solicitadas.

## 1. CARIMBO (canto inferior direito)
Extraia:
- Folha N°: número da prancha
- Unidade Construtiva: nome do sistema (ex: SISTEMA DE ESGOTAMENTO SANITÁRIO RONCADOR-PR)
- Título do projeto

## 2. RELAÇÃO DE MATERIAIS (geralmente canto superior direito)
Tabela com colunas: N° | MOS/Código | DISCRIMINAÇÃO/DESCRIÇÃO | QUANT | UN
Extraia TODOS os itens, inclusive os com quantidade "VB".

## 3. CLASSIFICAÇÃO
Classifique cada item em UMA das categorias:
{CATEGORIAS_LISTA}

Se nenhuma se encaixar, use "SERVIÇO".

## RESPOSTA
Retorne APENAS JSON válido (sem markdown):
{{
  "folha": "número ou null",
  "unidade_construtiva": "nome ou null",
  "titulo_projeto": "título ou null",
  "tem_relacao_materiais": true,
  "itens": [
    {{
      "numero": "N° do item",
      "codigo": "código MOS (pode ser vazio)",
      "descricao": "descrição completa",
      "quantidade": número ou null,
      "unidade": "UN/M/KG/VB etc",
      "categoria": "CATEGORIA DA LISTA"
    }}
  ]
}}

Se não houver tabela RELAÇÃO DE MATERIAIS, retorne tem_relacao_materiais: false e itens: [].
"""


# ---------------------------------------------------------------------------
# Extração
# ---------------------------------------------------------------------------

def render_page(pdf_path: str, page_num: int, dpi: int = 200) -> str:
    doc = fitz.open(pdf_path)
    try:
        page = doc.load_page(page_num)
        mat = fitz.Matrix(dpi / 72, dpi / 72)
        pix = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
        data = pix.tobytes("png")
    finally:
        doc.close()
    return base64.standard_b64encode(data).decode()


def _parse_json(text: str) -> dict | None:
    """Extrai JSON da resposta do Claude, lidando com blocos markdown."""
    # Remove blocos ```json ... ```
    clean = re.sub(r'```(?:json)?\s*', '', text).strip()
    # Tenta parsear direto
    try:
        return json.loads(clean)
    except json.JSONDecodeError:
        pass
    # Tenta extrair o maior bloco { ... }
    m = re.search(r'\{[\s\S]*\}', clean)
    if m:
        try:
            return json.loads(m.group())
        except json.JSONDecodeError:
            pass
    return None


def extract_from_image(client, img_b64: str, job_id: str = "", pg: int = 0, retries: int = 3) -> dict:
    img_bytes = base64.b64decode(img_b64)
    for attempt in range(retries):
        try:
            resp = client.models.generate_content(
                model="gemini-2.0-flash",
                contents=[
                    types.Part.from_bytes(data=img_bytes, mime_type="image/png"),
                    EXTRACTION_PROMPT,
                ],
            )
            text = resp.text.strip()
            result = _parse_json(text)
            if result:
                return result
            if job_id:
                _log(job_id, f"   ⚠️ Pág {pg+1}: resposta inesperada: {text[:200]}")
            return _empty()

        except Exception as e:
            err = str(e)
            if "429" in err or "quota" in err.lower() or "rate" in err.lower():
                wait = (attempt + 1) * 30
                if job_id:
                    _log(job_id, f"   ⏳ Rate limit. Aguardando {wait}s...")
                time.sleep(wait)
            else:
                if job_id:
                    _log(job_id, f"   ⚠️ Erro API (tentativa {attempt+1}): {e}")
                if attempt == retries - 1:
                    return _empty()
                time.sleep(5)

    return _empty()


def _empty() -> dict:
    return {"folha": None, "unidade_construtiva": None, "titulo_projeto": None,
            "tem_relacao_materiais": False, "itens": []}


def norm_qty(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace(" ", "").replace(",", "."))
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
THIN = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))

COLS = [
    ("Arquivo / Projeto Ref.", 30), ("Folha N°", 12),
    ("Unidade Construtiva", 35), ("Título do Projeto", 45),
    ("N° Item", 10), ("Código", 14), ("Descrição", 55),
    ("Quantidade", 14), ("Unidade", 10), ("Categoria", 25),
]
FIELDS = ["arquivo", "folha", "unidade_construtiva", "titulo_projeto",
          "numero_item", "codigo", "descricao", "quantidade", "unidade", "categoria"]


def save_excel(items: list, path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Base de Insumos"

    for ci, (h, w) in enumerate(COLS, 1):
        c = ws.cell(1, ci, h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.border = THIN
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    for ri, item in enumerate(items, 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, field in enumerate(FIELDS, 1):
            c = ws.cell(ri, ci, item.get(field))
            if fill: c.fill = fill
            c.border = THIN
            if field == "quantidade":
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif field in ("numero_item", "folha", "unidade", "codigo"):
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(vertical="center")

    # Aba Por Categoria
    ws2 = wb.create_sheet("Por Categoria")
    for ci, (h, w) in enumerate([("Categoria", 28), ("Itens", 12), ("% Total", 12)], 1):
        c = ws2.cell(1, ci, h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.border = THIN
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[1].height = 30

    counts: dict = {}
    for item in items:
        k = item.get("categoria") or "SERVIÇO"
        counts[k] = counts.get(k, 0) + 1
    total = len(items) or 1
    for ri, (cat, cnt) in enumerate(sorted(counts.items(), key=lambda x: -x[1]), 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, v in enumerate([cat, cnt, f"{cnt/total*100:.1f}%"], 1):
            c = ws2.cell(ri, ci, v)
            if fill: c.fill = fill
            c.border = THIN
            c.alignment = Alignment(horizontal="right" if ci > 1 else "left", vertical="center")

    # Aba Por Arquivo
    ws3 = wb.create_sheet("Por Arquivo")
    for ci, (h, w) in enumerate([("Arquivo", 40), ("Folhas", 20), ("Itens", 12)], 1):
        c = ws3.cell(1, ci, h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.border = THIN
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.row_dimensions[1].height = 30

    arqs: dict = {}
    for item in items:
        a = item["arquivo"]
        if a not in arqs:
            arqs[a] = {"folhas": set(), "cnt": 0}
        arqs[a]["folhas"].add(item["folha"])
        arqs[a]["cnt"] += 1
    for ri, (arq, info) in enumerate(sorted(arqs.items()), 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, v in enumerate([arq, ", ".join(sorted(info["folhas"])), info["cnt"]], 1):
            c = ws3.cell(ri, ci, v)
            if fill: c.fill = fill
            c.border = THIN
            c.alignment = Alignment(horizontal="right" if ci == 3 else "left", vertical="center")

    wb.save(path)


# ---------------------------------------------------------------------------
# Job processor (runs in background thread)
# ---------------------------------------------------------------------------

def _log(job_id: str, msg: str):
    with JOBS_LOCK:
        JOBS[job_id]["log"].append(msg)


def processar_job(job_id: str, pdf_files: list, api_key: str):
    client = genai.Client(api_key=api_key)
    job_dir = TEMP_DIR / job_id
    all_items = []

    try:
        for i, pdf_path in enumerate(pdf_files):
            with JOBS_LOCK:
                JOBS[job_id]["current"] = i + 1
                JOBS[job_id]["current_file"] = pdf_path.name

            _log(job_id, f"📄 [{i+1}/{len(pdf_files)}] {pdf_path.name}")

            doc = fitz.open(str(pdf_path))
            n_pages = len(doc)
            doc.close()

            for pg in range(n_pages):
                _log(job_id, f"   Página {pg+1}/{n_pages} — renderizando...")
                try:
                    img_b64 = render_page(str(pdf_path), pg)
                    _log(job_id, f"   Página {pg+1}/{n_pages} — extraindo com Gemini...")
                    data = extract_from_image(client, img_b64, job_id=job_id, pg=pg)
                    time.sleep(4)  # respeita limite de 15 req/min do plano gratuito
                except Exception as e:
                    _log(job_id, f"   ⚠️ Erro na pág {pg+1}: {e}")
                    continue

                if not data.get("tem_relacao_materiais") or not data.get("itens"):
                    _log(job_id, f"   Página {pg+1}/{n_pages} — sem RELAÇÃO DE MATERIAIS.")
                    continue

                folha = data.get("folha") or str(pg + 1)
                unidade = data.get("unidade_construtiva") or ""
                titulo = data.get("titulo_projeto") or ""
                count = 0

                for item in data["itens"]:
                    desc = (item.get("descricao") or "").strip()
                    if not desc:
                        continue
                    qtd = norm_qty(item.get("quantidade"))
                    cat = (item.get("categoria") or "SERVIÇO").strip().upper()
                    if cat not in CATEGORIAS:
                        match = next((c for c in CATEGORIAS if c.upper() == cat), None)
                        cat = match or "SERVIÇO"
                    all_items.append({
                        "arquivo": pdf_path.stem,
                        "folha": folha,
                        "unidade_construtiva": unidade,
                        "titulo_projeto": titulo,
                        "numero_item": str(item.get("numero") or ""),
                        "codigo": str(item.get("codigo") or ""),
                        "descricao": desc,
                        "quantidade": qtd,
                        "unidade": str(item.get("unidade") or ""),
                        "categoria": cat,
                    })
                    count += 1

                with JOBS_LOCK:
                    JOBS[job_id]["itens_extraidos"] = len(all_items)
                _log(job_id, f"   ✅ Página {pg+1} — {count} itens. (Folha {folha})")

        # Save Excel
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = str(job_dir / f"insumos_{ts}.xlsx")
        _log(job_id, f"💾 Gerando Excel com {len(all_items)} itens...")
        save_excel(all_items, excel_path)

        # Compute summary stats
        cats: dict = {}
        for item in all_items:
            k = item.get("categoria") or "SERVIÇO"
            cats[k] = cats.get(k, 0) + 1
        top_cats = sorted(cats.items(), key=lambda x: -x[1])[:8]

        with JOBS_LOCK:
            JOBS[job_id].update({
                "status": "concluido",
                "output": excel_path,
                "itens_extraidos": len(all_items),
                "top_categorias": top_cats,
                "total_arquivos_com_tabela": len({i["arquivo"] for i in all_items}),
            })
        _log(job_id, f"🎉 Concluído! {len(all_items)} itens extraídos.")

    except Exception as e:
        with JOBS_LOCK:
            JOBS[job_id].update({"status": "erro", "erro": str(e)})
        _log(job_id, f"❌ Erro fatal: {e}")


# ---------------------------------------------------------------------------
# Limpeza de jobs antigos
# ---------------------------------------------------------------------------

def cleanup_old_jobs():
    cutoff = datetime.now() - timedelta(hours=2)
    to_delete = []
    with JOBS_LOCK:
        for jid, job in JOBS.items():
            if job.get("created_at", datetime.now()) < cutoff:
                to_delete.append(jid)
    for jid in to_delete:
        jdir = TEMP_DIR / jid
        if jdir.exists():
            shutil.rmtree(jdir, ignore_errors=True)
        with JOBS_LOCK:
            JOBS.pop(jid, None)


# ---------------------------------------------------------------------------
# HTML
# ---------------------------------------------------------------------------

HTML = """<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Extrator de Insumos</title>
<style>
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  :root {
    --blue: #2563eb; --blue-dark: #1e40af; --blue-light: #eff6ff;
    --navy: #1e3a5f; --bg: #f1f5f9; --white: #ffffff;
    --text: #1e293b; --muted: #64748b; --border: #e2e8f0;
    --green: #16a34a; --green-light: #f0fdf4;
    --red: #dc2626; --red-light: #fef2f2;
    --yellow: #d97706; --yellow-light: #fffbeb;
    --radius: 12px; --shadow: 0 1px 3px rgba(0,0,0,.1), 0 4px 12px rgba(0,0,0,.06);
  }
  body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    background: var(--bg); color: var(--text); min-height: 100vh; }

  /* Header */
  header { background: var(--navy); color: #fff; padding: 0 32px;
    display: flex; align-items: center; gap: 16px; height: 64px;
    box-shadow: 0 2px 8px rgba(0,0,0,.25); }
  header .logo { font-size: 24px; }
  header h1 { font-size: 20px; font-weight: 700; letter-spacing: -.3px; }
  header span { font-size: 13px; color: #93c5fd; margin-left: 4px; }

  /* Layout */
  main { max-width: 900px; margin: 40px auto; padding: 0 24px 80px; }

  /* Card */
  .card { background: var(--white); border-radius: var(--radius);
    box-shadow: var(--shadow); padding: 28px 32px; margin-bottom: 24px; }
  .card-title { font-size: 15px; font-weight: 700; color: var(--navy);
    margin-bottom: 20px; display: flex; align-items: center; gap: 8px; }
  .card-title .icon { font-size: 18px; }

  /* API Key */
  #api-key-section label { display: block; font-size: 13px; font-weight: 600;
    color: var(--muted); margin-bottom: 8px; }
  #api-key-section input { width: 100%; padding: 10px 14px; border: 1.5px solid var(--border);
    border-radius: 8px; font-size: 14px; font-family: monospace;
    transition: border-color .15s; }
  #api-key-section input:focus { outline: none; border-color: var(--blue); }
  .hint { font-size: 12px; color: var(--muted); margin-top: 6px; }

  /* Drop zone */
  #drop-zone { border: 2.5px dashed var(--border); border-radius: 10px;
    padding: 48px 24px; text-align: center; cursor: pointer;
    transition: all .2s; background: #fafbfc; }
  #drop-zone:hover, #drop-zone.drag-over { border-color: var(--blue);
    background: var(--blue-light); }
  #drop-zone .dz-icon { font-size: 48px; margin-bottom: 12px; display: block; }
  #drop-zone .dz-text { font-size: 16px; font-weight: 600; color: var(--navy); }
  #drop-zone .dz-sub { font-size: 13px; color: var(--muted); margin-top: 4px; }
  #file-input { display: none; }

  /* File list */
  #file-list { margin-top: 16px; max-height: 220px; overflow-y: auto; }
  .file-item { display: flex; align-items: center; gap: 10px; padding: 8px 12px;
    border-radius: 8px; background: var(--bg); margin-bottom: 6px;
    font-size: 13px; }
  .file-item .fi-icon { font-size: 16px; flex-shrink: 0; }
  .file-item .fi-name { flex: 1; white-space: nowrap; overflow: hidden;
    text-overflow: ellipsis; color: var(--text); }
  .file-item .fi-size { color: var(--muted); font-size: 12px; flex-shrink: 0; }
  .file-item .fi-remove { cursor: pointer; color: var(--muted); font-size: 16px;
    padding: 2px 4px; border-radius: 4px; line-height: 1; }
  .file-item .fi-remove:hover { color: var(--red); background: var(--red-light); }

  /* Action bar */
  .action-bar { display: flex; align-items: center; justify-content: space-between;
    margin-top: 20px; gap: 12px; flex-wrap: wrap; }
  .file-count { font-size: 13px; color: var(--muted); }
  #btn-process { padding: 12px 28px; background: var(--blue); color: #fff;
    border: none; border-radius: 9px; font-size: 15px; font-weight: 700;
    cursor: pointer; transition: background .2s, transform .1s; }
  #btn-process:hover:not(:disabled) { background: var(--blue-dark); transform: translateY(-1px); }
  #btn-process:disabled { opacity: .5; cursor: not-allowed; transform: none; }
  #btn-clear { padding: 10px 18px; background: transparent; color: var(--muted);
    border: 1.5px solid var(--border); border-radius: 9px; font-size: 13px;
    cursor: pointer; transition: all .2s; }
  #btn-clear:hover { border-color: var(--red); color: var(--red); }

  /* Progress */
  #progress-section { display: none; }
  .progress-header { display: flex; align-items: center; justify-content: space-between;
    margin-bottom: 12px; }
  .progress-file { font-size: 14px; font-weight: 600; color: var(--navy);
    white-space: nowrap; overflow: hidden; text-overflow: ellipsis; max-width: 60%; }
  .progress-count { font-size: 13px; color: var(--muted); }
  .progress-bar-bg { background: var(--border); border-radius: 99px; height: 10px;
    overflow: hidden; margin-bottom: 16px; }
  .progress-bar-fill { height: 100%; background: linear-gradient(90deg, var(--blue), #60a5fa);
    border-radius: 99px; transition: width .4s ease; width: 0%;
    background-size: 200% 100%; animation: shimmer 1.5s infinite linear; }
  @keyframes shimmer {
    0% { background-position: 200% 0; }
    100% { background-position: -200% 0; }
  }
  .progress-stats { display: flex; gap: 16px; margin-bottom: 16px; flex-wrap: wrap; }
  .stat-chip { background: var(--blue-light); color: var(--blue-dark); padding: 6px 14px;
    border-radius: 99px; font-size: 13px; font-weight: 600; }
  #log-box { background: #0f172a; border-radius: 8px; padding: 14px 16px;
    max-height: 200px; overflow-y: auto; font-family: monospace; font-size: 12px;
    color: #94a3b8; line-height: 1.7; }
  #log-box .log-line { animation: fadein .3s ease; }
  @keyframes fadein { from { opacity: 0; transform: translateY(4px); } to { opacity: 1; } }

  /* Results */
  #results-section { display: none; }
  .result-banner { background: var(--green-light); border: 1.5px solid #bbf7d0;
    border-radius: 10px; padding: 20px 24px; display: flex; align-items: center;
    gap: 16px; margin-bottom: 20px; flex-wrap: wrap; }
  .result-banner .rb-icon { font-size: 36px; }
  .result-banner .rb-text h3 { font-size: 18px; font-weight: 700; color: var(--green); }
  .result-banner .rb-text p { font-size: 14px; color: var(--muted); margin-top: 2px; }
  #btn-download { padding: 13px 32px; background: var(--green); color: #fff;
    border: none; border-radius: 9px; font-size: 15px; font-weight: 700;
    cursor: pointer; transition: background .2s; text-decoration: none;
    display: inline-block; }
  #btn-download:hover { background: #15803d; }

  .cats-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr));
    gap: 10px; margin-top: 16px; }
  .cat-card { background: var(--bg); border-radius: 8px; padding: 12px 14px;
    display: flex; justify-content: space-between; align-items: center; gap: 8px; }
  .cat-card .cat-name { font-size: 12px; font-weight: 600; color: var(--text);
    white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
  .cat-card .cat-count { background: var(--blue); color: #fff; border-radius: 99px;
    padding: 2px 10px; font-size: 12px; font-weight: 700; flex-shrink: 0; }

  /* Error */
  #error-section { display: none; background: var(--red-light); border: 1.5px solid #fecaca;
    border-radius: var(--radius); padding: 20px 24px; }
  #error-section h3 { color: var(--red); font-size: 16px; font-weight: 700; margin-bottom: 8px; }
  #error-section p { font-size: 14px; color: var(--text); }

  /* Key OK badge */
  .key-ok { display: inline-flex; align-items: center; gap: 6px; background: var(--green-light);
    color: var(--green); border-radius: 8px; padding: 8px 14px; font-size: 13px;
    font-weight: 600; }

  /* Reset link */
  #btn-reset { display: none; margin-top: 12px; background: none; border: none;
    color: var(--blue); font-size: 13px; cursor: pointer; text-decoration: underline; }
</style>
</head>
<body>

<header>
  <span class="logo">⚙️</span>
  <h1>Extrator de Insumos</h1>
  <span>Pranchas de Projeto → Base Excel</span>
</header>

<main>

  <!-- API Key -->
  <div class="card" id="api-key-section">
    <div class="card-title"><span class="icon">🔑</span> Chave da API Google Gemini</div>
    <div id="key-env-ok" style="display:none" class="key-ok">✅ GEMINI_API_KEY configurada via variável de ambiente.</div>
    <div id="key-input-area">
      <label for="api-key-input">Cole sua chave API do Google Gemini (AIza...)</label>
      <input type="password" id="api-key-input" placeholder="AIzaSy..." autocomplete="off">
      <p class="hint">Chave gratuita em <strong>aistudio.google.com</strong> → Get API Key. Usada apenas nesta sessão, nunca armazenada.</p>
    </div>
  </div>

  <!-- Upload -->
  <div class="card">
    <div class="card-title"><span class="icon">📂</span> Selecionar PDFs</div>
    <div id="drop-zone">
      <span class="dz-icon">📄</span>
      <div class="dz-text">Arraste os PDFs aqui</div>
      <div class="dz-sub">ou clique para selecionar arquivos</div>
    </div>
    <input type="file" id="file-input" accept=".pdf" multiple>
    <div id="file-list"></div>
    <div class="action-bar">
      <span class="file-count" id="file-count">Nenhum arquivo selecionado</span>
      <div style="display:flex;gap:10px">
        <button id="btn-clear" onclick="clearFiles()">Limpar</button>
        <button id="btn-process" disabled onclick="startProcessing()">▶ Extrair Insumos</button>
      </div>
    </div>
  </div>

  <!-- Progress -->
  <div class="card" id="progress-section">
    <div class="card-title"><span class="icon">⚙️</span> Processando...</div>
    <div class="progress-header">
      <div class="progress-file" id="progress-file">Iniciando...</div>
      <div class="progress-count" id="progress-count">0 / 0</div>
    </div>
    <div class="progress-bar-bg"><div class="progress-bar-fill" id="progress-bar"></div></div>
    <div class="progress-stats">
      <div class="stat-chip" id="stat-itens">0 itens extraídos</div>
    </div>
    <div id="log-box"></div>
  </div>

  <!-- Results -->
  <div class="card" id="results-section">
    <div class="card-title"><span class="icon">✅</span> Extração Concluída</div>
    <div class="result-banner">
      <span class="rb-icon">📊</span>
      <div class="rb-text">
        <h3 id="result-title">Base de insumos pronta!</h3>
        <p id="result-sub"></p>
      </div>
      <a id="btn-download" href="#" download>⬇ Baixar Excel</a>
    </div>
    <div class="card-title" style="margin-bottom:8px"><span class="icon">📋</span> Top Categorias</div>
    <div class="cats-grid" id="cats-grid"></div>
    <button id="btn-reset" onclick="resetApp()">⟳ Processar novos arquivos</button>
  </div>

  <!-- Error -->
  <div id="error-section">
    <h3>❌ Erro no processamento</h3>
    <p id="error-msg"></p>
    <button id="btn-reset-err" onclick="resetApp()" style="margin-top:16px;padding:8px 18px;background:var(--red);color:#fff;border:none;border-radius:8px;cursor:pointer;font-size:14px;">Tentar novamente</button>
  </div>

</main>

<script>
let selectedFiles = [];
let currentJobId = null;
let pollTimer = null;

// ── Check API key status on load ──────────────────────────────────────────
fetch('/api-key-status').then(r => r.json()).then(d => {
  if (d.configured) {
    document.getElementById('key-env-ok').style.display = 'flex';
    document.getElementById('key-input-area').style.display = 'none';
  }
});

// ── Drop zone ──────────────────────────────────────────────────────────────
const dropZone = document.getElementById('drop-zone');
const fileInput = document.getElementById('file-input');

dropZone.addEventListener('click', () => fileInput.click());
fileInput.addEventListener('change', () => addFiles(fileInput.files));

dropZone.addEventListener('dragover', e => { e.preventDefault(); dropZone.classList.add('drag-over'); });
dropZone.addEventListener('dragleave', () => dropZone.classList.remove('drag-over'));
dropZone.addEventListener('drop', e => {
  e.preventDefault();
  dropZone.classList.remove('drag-over');
  addFiles(e.dataTransfer.files);
});

function addFiles(fileList) {
  const existing = new Set(selectedFiles.map(f => f.name + f.size));
  for (const f of fileList) {
    if (f.name.toLowerCase().endsWith('.pdf') && !existing.has(f.name + f.size)) {
      selectedFiles.push(f);
      existing.add(f.name + f.size);
    }
  }
  renderFileList();
}

function removeFile(idx) {
  selectedFiles.splice(idx, 1);
  renderFileList();
}

function clearFiles() {
  selectedFiles = [];
  renderFileList();
  fileInput.value = '';
}

function fmtSize(bytes) {
  if (bytes < 1024) return bytes + ' B';
  if (bytes < 1024 * 1024) return (bytes / 1024).toFixed(1) + ' KB';
  return (bytes / (1024 * 1024)).toFixed(1) + ' MB';
}

function renderFileList() {
  const list = document.getElementById('file-list');
  const count = document.getElementById('file-count');
  const btn = document.getElementById('btn-process');

  if (selectedFiles.length === 0) {
    list.innerHTML = '';
    count.textContent = 'Nenhum arquivo selecionado';
    btn.disabled = true;
    return;
  }

  count.textContent = selectedFiles.length + ' arquivo(s) selecionado(s)';
  btn.disabled = false;

  list.innerHTML = selectedFiles.map((f, i) => `
    <div class="file-item">
      <span class="fi-icon">📄</span>
      <span class="fi-name" title="${f.name}">${f.name}</span>
      <span class="fi-size">${fmtSize(f.size)}</span>
      <span class="fi-remove" onclick="removeFile(${i})" title="Remover">✕</span>
    </div>
  `).join('');
}

// ── Processing ─────────────────────────────────────────────────────────────
async function startProcessing() {
  const apiKey = document.getElementById('api-key-input').value.trim()
    || null;

  if (!apiKey) {
    const keyStatus = await fetch('/api-key-status').then(r => r.json());
    if (!keyStatus.configured) {
      alert('Informe a chave da API Google Gemini antes de continuar.');
      return;
    }
  }

  if (selectedFiles.length === 0) return;

  // Show progress, hide others
  document.getElementById('progress-section').style.display = 'block';
  document.getElementById('results-section').style.display = 'none';
  document.getElementById('error-section').style.display = 'none';
  document.getElementById('btn-process').disabled = true;
  document.getElementById('log-box').innerHTML = '';

  const formData = new FormData();
  for (const f of selectedFiles) formData.append('pdfs', f);
  if (apiKey) formData.append('api_key', apiKey);

  try {
    const resp = await fetch('/processar', { method: 'POST', body: formData });
    const data = await resp.json();
    if (!resp.ok || data.erro) {
      showError(data.erro || 'Falha ao iniciar processamento.');
      return;
    }
    currentJobId = data.job_id;
    pollTimer = setInterval(pollStatus, 1200);
  } catch (e) {
    showError('Erro de conexão: ' + e.message);
  }
}

let lastLogCount = 0;

async function pollStatus() {
  try {
    const resp = await fetch('/status/' + currentJobId);
    const d = await resp.json();

    // Update progress
    document.getElementById('progress-file').textContent = d.current_file || 'Processando...';
    document.getElementById('progress-count').textContent = d.current + ' / ' + d.total;
    document.getElementById('stat-itens').textContent = d.itens_extraidos + ' itens extraídos';
    const pct = d.total > 0 ? Math.round(d.current / d.total * 100) : 0;
    document.getElementById('progress-bar').style.width = pct + '%';

    // Append new log lines
    const logBox = document.getElementById('log-box');
    const lines = d.log || [];
    for (let i = lastLogCount; i < lines.length; i++) {
      const div = document.createElement('div');
      div.className = 'log-line';
      div.textContent = lines[i];
      logBox.appendChild(div);
    }
    lastLogCount = lines.length;
    logBox.scrollTop = logBox.scrollHeight;

    if (d.status === 'concluido') {
      clearInterval(pollTimer);
      document.getElementById('progress-bar').style.width = '100%';
      showResults(d);
    } else if (d.status === 'erro') {
      clearInterval(pollTimer);
      showError(d.erro || 'Erro desconhecido');
    }
  } catch (e) {
    // network hiccup, keep polling
  }
}

function showResults(d) {
  document.getElementById('results-section').style.display = 'block';
  document.getElementById('btn-reset').style.display = 'inline-block';

  document.getElementById('result-title').textContent =
    d.itens_extraidos + ' itens extraídos com sucesso!';
  document.getElementById('result-sub').textContent =
    d.total_arquivos_com_tabela + ' arquivo(s) continham tabela RELAÇÃO DE MATERIAIS';

  const dlBtn = document.getElementById('btn-download');
  dlBtn.href = '/download/' + currentJobId;

  const grid = document.getElementById('cats-grid');
  grid.innerHTML = (d.top_categorias || []).map(([cat, cnt]) => `
    <div class="cat-card">
      <span class="cat-name" title="${cat}">${cat}</span>
      <span class="cat-count">${cnt}</span>
    </div>
  `).join('');
}

function showError(msg) {
  document.getElementById('error-section').style.display = 'block';
  document.getElementById('error-msg').textContent = msg;
  document.getElementById('btn-process').disabled = false;
}

function resetApp() {
  clearFiles();
  currentJobId = null;
  lastLogCount = 0;
  if (pollTimer) clearInterval(pollTimer);
  document.getElementById('progress-section').style.display = 'none';
  document.getElementById('results-section').style.display = 'none';
  document.getElementById('error-section').style.display = 'none';
  document.getElementById('btn-reset').style.display = 'none';
  document.getElementById('btn-process').disabled = true;
  document.getElementById('log-box').innerHTML = '';
  document.getElementById('progress-bar').style.width = '0%';
}
</script>
</body>
</html>
"""


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return HTML


@app.route("/api-key-status")
def api_key_status():
    return jsonify({"configured": bool(os.environ.get("GEMINI_API_KEY"))})


@app.route("/processar", methods=["POST"])
def processar():
    cleanup_old_jobs()

    api_key = (
        os.environ.get("GEMINI_API_KEY")
        or request.form.get("api_key", "").strip()
    )
    if not api_key:
        return jsonify({"erro": "GEMINI_API_KEY não configurada"}), 400

    files = request.files.getlist("pdfs")
    if not files:
        return jsonify({"erro": "Nenhum arquivo enviado"}), 400

    job_id = str(uuid.uuid4())[:8]
    job_dir = TEMP_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)

    saved = []
    for f in files:
        if f.filename and f.filename.lower().endswith(".pdf"):
            fname = secure_filename(f.filename)
            fpath = job_dir / fname
            f.save(str(fpath))
            saved.append(fpath)

    if not saved:
        return jsonify({"erro": "Nenhum PDF válido recebido"}), 400

    with JOBS_LOCK:
        JOBS[job_id] = {
            "status": "processando",
            "total": len(saved),
            "current": 0,
            "current_file": "",
            "itens_extraidos": 0,
            "log": [],
            "output": None,
            "erro": None,
            "top_categorias": [],
            "total_arquivos_com_tabela": 0,
            "created_at": datetime.now(),
        }

    thread = threading.Thread(
        target=processar_job, args=(job_id, saved, api_key), daemon=True
    )
    thread.start()
    return jsonify({"job_id": job_id})


@app.route("/status/<job_id>")
def status(job_id):
    with JOBS_LOCK:
        job = JOBS.get(job_id)
    if not job:
        return jsonify({"erro": "Job não encontrado"}), 404
    return jsonify({
        "status": job["status"],
        "total": job["total"],
        "current": job["current"],
        "current_file": job["current_file"],
        "itens_extraidos": job["itens_extraidos"],
        "log": job["log"][-30:],
        "erro": job["erro"],
        "top_categorias": job.get("top_categorias", []),
        "total_arquivos_com_tabela": job.get("total_arquivos_com_tabela", 0),
    })


@app.route("/download/<job_id>")
def download(job_id):
    with JOBS_LOCK:
        job = JOBS.get(job_id)
    if not job or job["status"] != "concluido" or not job["output"]:
        return jsonify({"erro": "Arquivo não disponível"}), 404
    path = Path(job["output"])
    if not path.exists():
        return jsonify({"erro": "Arquivo não encontrado no servidor"}), 404
    return send_file(str(path), as_attachment=True, download_name=path.name)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    key_set = bool(os.environ.get("GEMINI_API_KEY"))
    print("=" * 55)
    print("  Extrator de Insumos - Iniciando servidor")
    print("=" * 55)
    print(f"  URL:        http://localhost:{port}")
    print(f"  API Key:    {'✓ GEMINI_API_KEY configurada' if key_set else '⚠ não definida (informe no navegador)'}")
    print(f"  Temp dir:   {TEMP_DIR}")
    print("=" * 55)
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
